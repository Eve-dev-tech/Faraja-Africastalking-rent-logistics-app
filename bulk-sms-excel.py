#importing necessary libraries
import os
import africastalking as at
from openpyxl import load_workbook
from dotenv import load_dotenv

# the load_dotenv function gets the environment variables defined in .env file
load_dotenv()

# assigns the variables to the environment variables
#api_key = os.getenv("api_key")
#username = os.getenv("username")

username = "enter your user name here"
api_key = "Enter your api_key here"
# Initialize the Africas Talking client with the required credentials
at.initialize(username, api_key)
sms = at.SMS

# create a function to parse the excel and send a customized message
wb = load_workbook('sample.xlsx')
print(wb.sheetnames)
sheet1 = wb['Sheet1']

names_cell_range = sheet1['B2:B4']
number_cell_range = sheet1['C2:C4']


def send_messages():
    for row in sheet1.iter_rows(values_only=True):
        name = row[1]
        number = f"+254{row[2]}"
        house_number = row[3]
        rent_due_date = "Friday 14 July at 8.00 am "
        print(name,number)
        message = f"hey {name}  Kindly note your rent for your house number {house_number}  is due on {rent_due_date} dial *384*87779# to make payment or report issue "
        try:
            response = sms.send(message, [number])
            print(response)
        except Exception as e:
            print(f"Imagine kamekuramba!!: {e}")


send_messages()

